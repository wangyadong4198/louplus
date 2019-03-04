# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

def combine():
    student_dic = {}
    time_dic = {}
    stime_dic = {}
    wb = Workbook()
    wb = load_workbook('courses.xlsx')
    cells = wb['students']
    for i in range(1,486):
        key = cells['B{}'.format(i)].value
        value = cells['C{}'.format(i)].value
        value2 = cells['A{}'.format(i)].value
        student_dic[key] = value
        stime_dic[key] = value2
    cells = wb['time']
    for i in range(1,486):
        key = cells['B{}'.format(i)].value
        value = cells['C{}'.format(i)].value
        time_dic[key] = value
    ws = wb.create_sheet(title = 'combine')
    row = 1
    for key in student_dic.keys(): 
        ws.cell(column=1,row=row,value=stime_dic[key])
        ws.cell(column=2,row=row,value=key)
        ws.cell(column=3,row=row,value=student_dic[key])
        ws.cell(column=4,row=row,value=time_dic[key])
        row += 1
    wb.save(filename='courses.xlsx')

def split():
    wb = Workbook()
    wb = load_workbook('courses.xlsx')
    cells = wb['combine']
    years = []
    for i in range(2,486):
        stime = cells['A{}'.format(i)].value.year
        if stime in years:
            pass
        else:
            years.append(stime)
    for year in years:
        print('year.....:{}'.format(year))
        wb2 = Workbook()
        wb2.remove(wb2.active)
        ws = wb2.create_sheet(title = '{}'.format(year))
        ws.cell(column=1,row=1,value=cells['A1'].value)
        ws.cell(column=2,row=1,value=cells['B1'].value)
        ws.cell(column=3,row=1,value=cells['C1'].value)
        ws.cell(column=4,row=1,value=cells['D1'].value)
        row = 2
        for i in range(2,486):
            if cells['A{}'.format(i)].value.year == year:
               # print(cells['A{}'.format(i)].value)
                ws.cell(column=1,row=row,value=cells['A{}'.format(i)].value)
                ws.cell(column=2,row=row,value=cells['B{}'.format(i)].value)
                ws.cell(column=3,row=row,value=cells['C{}'.format(i)].value)
                ws.cell(column=4,row=row,value=cells['D{}'.format(i)].value)
                row += 1
        wb2.save(filename='{}.xlsx'.format(year))
    
    
if __name__ == '__main__':
    combine()
    split()
